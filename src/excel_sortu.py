"""
emaitzak/ karpetako JSON emaitzak Excel fitxategietara bihurtu.
Talde bakoitzeko Excel bat sortzen du: Notak, Feedbacka, Akatsak orriekin.
"""

import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

EMAITZAK_DIR = Path(__file__).resolve().parent.parent / "emaitzak"

# Talde-konfigurazioa: (karpeta, azpikarpeta, output)
TALDEAK = {
    "DBH4A": "Idazlan-koadernoa (entregatzea)",
    "DBH4B": "01_narrazioa",
    "DBH3A": "idazlana",
    "DBH3B": "idazlana",
}

# Estiloak
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
IA_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def lortu_timestamp(fitxategi_izena):
    """Atera timestamp-a fitxategi-izenetik: _YYYYMMDD_HHMMSS.json"""
    m = re.search(r"_(\d{8}_\d{6})\.json$", fitxategi_izena)
    return m.group(1) if m else "000000_000000"


def lortu_ikasle_izena(fitxategi_izena):
    """Atera ikaslearen izena fitxategi-izenetik (timestamp kendu)."""
    izena = re.sub(r"_\d{8}_\d{6}\.json$", "", fitxategi_izena)
    return izena.strip()


def normalizatu_ikasle_gakoa(fitxategi_izena):
    """Fitxategi-izena normalizatu ikasle berdinak taldekatzeko.
    'Koadernoa(1)', 'koadernoa', 'fitxategiaren kopia', etc. berdindu."""
    izena = lortu_ikasle_izena(fitxategi_izena)
    # Kendu " - Idazlan..." eta ondorengoa
    izena = re.sub(r"\s*-\s*[Ii]dazlan.*$", "", izena)
    izena = re.sub(r"\s*-\s*IDAZLAN.*$", "", izena)
    izena = re.sub(r"\s*-\s*[Ee]uskara.*$", "", izena)
    izena = re.sub(r"\s*-\s*BIZITZAREN.*$", "", izena)
    izena = re.sub(r"\s*-\s*XXV\..*$", "", izena)
    izena = re.sub(r"\s*fitxategiaren kopia.*$", "", izena, flags=re.IGNORECASE)
    izena = re.sub(r"\(\d+\)\s*$", "", izena)
    izena = re.sub(r"_+$", "", izena)
    return izena.strip().lower()


def lortu_azken_jsonak(karpeta):
    """Karpeta bateko JSON fitxategiak bildu, ikasle bakoitzeko azkenekoa soilik."""
    if not karpeta.exists():
        return []

    fitxategiak = defaultdict(list)
    for f in karpeta.glob("*.json"):
        gakoa = normalizatu_ikasle_gakoa(f.name)
        fitxategiak[gakoa].append(f)

    # Ikasle bakoitzeko timestamp handiena (azkenekoa) aukeratu
    azkenak = []
    for gakoa, paths in fitxategiak.items():
        paths.sort(key=lambda p: lortu_timestamp(p.name))
        azkenak.append(paths[-1])  # azkenekoa = berriena

    return sorted(azkenak, key=lambda p: normalizatu_ikasle_gakoa(p.name))


def kargatu_json(path):
    """JSON fitxategia kargatu."""
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def lortu_ia_info(data):
    """IA segurtasun-azterketa informazioa atera, formatu desberdinak kudeatuz."""
    # Lehentasuna: meta.segurtasun_azterketa
    meta_sa = data.get("meta", {}).get("segurtasun_azterketa", {})
    if meta_sa:
        return meta_sa

    # Fallback: top-level segurtasun_azterketa
    top_sa = data.get("segurtasun_azterketa", {})
    if top_sa:
        # Bihurtu formatu estandarrera
        seinaleak = top_sa.get("seinaleak", {})
        if isinstance(seinaleak, list):
            # "S1: 2/3 - ..." formatutik dict-era
            seinale_dict = {}
            for s in seinaleak:
                m = re.match(r"(S\d+):\s*(\d+)", s)
                if m:
                    seinale_dict[m.group(1)] = int(m.group(2))
            top_sa = dict(top_sa)
            top_sa["seinaleak"] = seinale_dict
        return top_sa

    return {}


def da_ia_blokeatua(data):
    """Ikaslea IA dela susmatzeagatik blokeatuta dagoen.
    Blokeatu = ebaluazioa hutsa + maila ALTUA."""
    eb = data.get("ebaluazioa", {})
    if eb:  # ebaluazioak balioak baditu, ez dago blokeatuta
        return False
    ia = lortu_ia_info(data)
    return ia.get("maila") == "ALTUA"


def lortu_ia_justifikazioa(data, ikasle_izena):
    """IA blokeatuen justifikazioa sortu: S1-S7 balioak + iruzkina."""
    ia = lortu_ia_info(data)
    seinaleak = ia.get("seinaleak", {})
    iruzkina = ia.get("iruzkina", "")
    puntuazioa = ia.get("ia_puntuazioa", "?")

    # S1-S7 lerroak
    seinale_deskribapenak = {
        "S1": "Perfekzio linguistikoa",
        "S2": "Hiztegi sofistikatua",
        "S3": "Egitura perfektua",
        "S4": "Tonu robotikoa",
        "S5": "Luzera handia",
        "S6": "Akats natural falta",
        "S7": "Kalitate uniformea",
    }

    lerroak = [f"IA puntuazioa: {puntuazioa}/21\n"]
    for skey in ["S1", "S2", "S3", "S4", "S5", "S6", "S7"]:
        balioa = seinaleak.get(skey, "?")
        desk = seinale_deskribapenak.get(skey, "")
        lerroak.append(f"{skey}={balioa}/3 ({desk})")

    lerroak.append(f"\n{iruzkina}")
    return "\n".join(lerroak)


def lortu_ikasle_izena_json(data, fitxategi_path):
    """Ikasle izena lortu: JSON kodea edo fitxategi izena."""
    kodea = data.get("kodea", "")
    # Fitxategi izenetik izena atera (osoagoa da normalean)
    izena = lortu_ikasle_izena(fitxategi_path.name)
    # Kendu "- Idazlan Koadernoa" eta antzekoak
    izena = re.sub(r"\s*-\s*[Ii]dazlan.*$", "", izena)
    izena = re.sub(r"\s*-\s*IDAZLAN.*$", "", izena)
    izena = re.sub(r"\s*-\s*Euskara.*$", "", izena)
    izena = re.sub(r"\s*-\s*BIZITZAREN.*$", "", izena)
    izena = re.sub(r"\s*fitxategiaren kopia$", "", izena, flags=re.IGNORECASE)
    izena = re.sub(r"\(\d+\)$", "", izena)  # (1) kendu
    izena = re.sub(r"__$", "", izena)
    izena = izena.strip()
    return izena if izena else kodea


def estilotu_header(ws, zutabe_kop):
    """Header errenkada estilotu."""
    for col in range(1, zutabe_kop + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autofit_zutabeak(ws, min_width=12, max_width=60):
    """Zutabe-zabalera automatikoki doitu."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_line = max(len(line) for line in lines)
                max_len = max(max_len, max_line)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def sortu_notak_orria(wb, datuak):
    """Notak orria sortu."""
    ws = wb.active
    ws.title = "Notak"

    headerrak = [
        "Ikaslea", "AB", "AK", "BLH", "ZL",
        "Nota finala", "Heziberri maila", "EP100",
        "IA maila", "IA oharra", "IA justifikazioa"
    ]
    ws.append(headerrak)
    estilotu_header(ws, len(headerrak))

    for izena, data, path in datuak:
        blokeatua = da_ia_blokeatua(data)
        ia = lortu_ia_info(data)
        ia_maila = ia.get("maila", "BAXUA")

        if blokeatua:
            row = [
                izena, "-", "-", "-", "-",
                "-", "-", "-",
                ia_maila,
                "IA susmo altua",
                lortu_ia_justifikazioa(data, izena),
            ]
        else:
            eb = data.get("ebaluazioa", {})
            lab = data.get("laburpena", {})
            ep = data.get("EP100", {})

            ab = eb.get("atazaren_betetzea", {}).get("nota", "")
            ak = eb.get("antolaketa_koherentzia", {}).get("nota", "")
            blh = eb.get("baliabide_linguistikoen_hedadura", {}).get("nota", "")
            zl = eb.get("zuzentasun_linguistikoa", {}).get("nota", "")
            nota = lab.get("biribiltzea", "")
            maila = lab.get("heziberri_maila", "")
            ep100 = ep.get("EP100_pisuduna", "")

            row = [
                izena, ab, ak, blh, zl,
                nota, maila, ep100,
                ia_maila, "", "",
            ]

        ws.append(row)
        row_idx = ws.max_row

        # IA blokeatuen errenkadak gorriz markatu
        if blokeatua:
            for col in range(1, len(headerrak) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = IA_FILL
                cell.border = THIN_BORDER
        else:
            for col in range(1, len(headerrak) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # IA justifikazioa zutabea wrap_text
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.alignment = WRAP

    autofit_zutabeak(ws)
    # Justifikazioa zutabea zabalago
    ws.column_dimensions["K"].width = 80

    # Izoztatu headerrak
    ws.freeze_panes = "A2"


def sortu_feedbacka_orria(wb, datuak):
    """Feedbacka orria sortu: irizpide bakoitzeko iruzkina osoa."""
    ws = wb.create_sheet("Feedbacka")

    irizpideak = [
        ("atazaren_betetzea", "AB"),
        ("antolaketa_koherentzia", "AK"),
        ("baliabide_linguistikoen_hedadura", "BLH"),
        ("zuzentasun_linguistikoa", "ZL"),
    ]

    headerrak = ["Ikaslea"]
    for _, laburdura in irizpideak:
        headerrak.extend([
            f"{laburdura} - Indarguneak",
            f"{laburdura} - Hobetzekoak",
            f"{laburdura} - Feed forward",
        ])
    headerrak.append("Laburpen orokorra")
    headerrak.append("Lehentasuna")

    ws.append(headerrak)
    estilotu_header(ws, len(headerrak))

    for izena, data, path in datuak:
        blokeatua = da_ia_blokeatua(data)
        eb = data.get("ebaluazioa", {})
        lab = data.get("laburpena", {})

        row = [izena]

        if blokeatua:
            for _ in irizpideak:
                row.extend(["-", "-", "-"])
            row.append(lab.get("orokorra", "IA susmo altua"))
            row.append("-")
        else:
            for gakoa, _ in irizpideak:
                irizpidea = eb.get(gakoa, {})
                iruzkina = irizpidea.get("iruzkina", {})
                indarguneak = iruzkina.get("indarguneak", [])
                hobetzekoak = iruzkina.get("hobetzekoak", [])
                feed_forward = iruzkina.get("feed_forward", "")

                row.append("\n".join(f"- {i}" for i in indarguneak) if indarguneak else "")
                row.append("\n".join(f"- {h}" for h in hobetzekoak) if hobetzekoak else "")
                row.append(feed_forward)

            row.append(lab.get("orokorra", ""))
            row.append(lab.get("lehentasuna", ""))

        ws.append(row)
        row_idx = ws.max_row

        if blokeatua:
            for col in range(1, len(headerrak) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = IA_FILL
                cell.border = THIN_BORDER

        for col in range(1, len(headerrak) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    autofit_zutabeak(ws, max_width=50)
    ws.freeze_panes = "B2"


def sortu_akatsak_orria(wb, datuak):
    """Akatsak orria sortu."""
    ws = wb.create_sheet("Akatsak")

    headerrak = [
        "Ikaslea", "Kategoria", "Akats ID", "Kopurua",
        "Larritasuna", "Detektatutakoa", "Zuzenketa", "Azalpena",
    ]
    ws.append(headerrak)
    estilotu_header(ws, len(headerrak))

    for izena, data, path in datuak:
        blokeatua = da_ia_blokeatua(data)
        akatsak = data.get("akats_taldeak", [])

        if blokeatua:
            row = [izena, "IA susmo altua", "-", "-", "-", "-", "-", "-"]
            ws.append(row)
            row_idx = ws.max_row
            for col in range(1, len(headerrak) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = IA_FILL
                cell.border = THIN_BORDER
        elif not akatsak:
            row = [izena, "(akatsik ez)", "-", "-", "-", "-", "-", "-"]
            ws.append(row)
            row_idx = ws.max_row
            for col in range(1, len(headerrak) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
        else:
            for i, akatsa in enumerate(akatsak):
                adibidea = akatsa.get("adibidea", {})
                row = [
                    izena if i == 0 else "",
                    akatsa.get("kategoria", akatsa.get("categoria", "")),
                    akatsa.get("akats_id", ""),
                    akatsa.get("kopurua", ""),
                    akatsa.get("larritasuna", ""),
                    adibidea.get("detektatutakoa", ""),
                    adibidea.get("zuzenketa", ""),
                    adibidea.get("azalpena", ""),
                ]
                ws.append(row)
                row_idx = ws.max_row
                for col in range(1, len(headerrak) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.alignment = WRAP
                    cell.border = THIN_BORDER

    autofit_zutabeak(ws, max_width=50)
    ws.freeze_panes = "B2"


def sortu_talde_excel(taldea, azpikarpeta):
    """Talde baten Excel fitxategia sortu."""
    karpeta = EMAITZAK_DIR / taldea / azpikarpeta

    if not karpeta.exists():
        print(f"  [!] Karpeta ez da existitzen: {karpeta}")
        return None

    json_fitxategiak = lortu_azken_jsonak(karpeta)
    if not json_fitxategiak:
        print(f"  [!] JSON fitxategirik ez: {karpeta}")
        return None

    # Datuak kargatu
    datuak = []
    for path in json_fitxategiak:
        data = kargatu_json(path)
        izena = lortu_ikasle_izena_json(data, path)
        datuak.append((izena, data, path))

    # Alfabetikoki ordenatu
    datuak.sort(key=lambda x: x[0].lower())

    print(f"  {len(datuak)} ikasle aurkitu")

    # Workbook sortu
    wb = Workbook()
    sortu_notak_orria(wb, datuak)
    sortu_feedbacka_orria(wb, datuak)
    sortu_akatsak_orria(wb, datuak)

    # Gorde
    output_path = EMAITZAK_DIR / f"{taldea}_emaitzak.xlsx"
    wb.save(output_path)
    print(f"  -> {output_path}")

    # Estatistikak
    blokeatuak = sum(1 for _, d, _ in datuak if da_ia_blokeatua(d))
    normalak = len(datuak) - blokeatuak
    print(f"     {normalak} normal + {blokeatuak} IA blokeatuta")

    return output_path


def main():
    print("=" * 60)
    print("ZUZENDU — Excel emaitzak sortzailea")
    print("=" * 60)

    sortutakoak = []
    for taldea, azpikarpeta in TALDEAK.items():
        print(f"\n[{taldea}]")
        result = sortu_talde_excel(taldea, azpikarpeta)
        if result:
            sortutakoak.append(result)

    print(f"\n{'=' * 60}")
    print(f"Guztira {len(sortutakoak)} Excel fitxategi sortu dira.")
    print("=" * 60)


if __name__ == "__main__":
    main()
