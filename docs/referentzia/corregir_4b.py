# -*- coding: utf-8 -*-
import json, re
from pathlib import Path
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic

import os
from dotenv import load_dotenv
load_dotenv()
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

carpetas = [
    Path(r'C:\Users\usuario\Downloads\Idazlan koadernoa-20260208T213335Z-1-001\Idazlan koadernoa'),
    Path(r'H:\Mi unidad\Classroom\4 B\Idazlan-koadernoa (entregatzea)')
]

print('=== 4B - IDAZLAN KOADERNOA ===')

resultados = {}

for carpeta in carpetas:
    archivos = list(carpeta.glob('*.docx'))
    print(f'\nCarpeta: {carpeta.name} ({len(archivos)} archivos)')

    for f in archivos:
        if 'Template' in f.name or 'XXV.' in f.name or 'Euskara Idazlan' in f.name:
            continue
        nombre = f.stem.split(' - ')[0].strip()

        if nombre in resultados:
            continue

        print(f'  {nombre}...', end=' ', flush=True)

        try:
            doc = Document(f)
            texto = ' '.join([p.text for p in doc.paragraphs if p.text.strip()])
            if len(texto) < 50:
                print('SKIP')
                continue

            # PASO 1: Detectar IA
            prompt_ia = f"""Aztertu testu hau DBH 4. mailako ikasle batek idatzita dagoen edo Adimen Artifizialak sortua den.

IA-REN ADIERAZLEAK:
- Koherentzia gehiegizkoa (B1 mailako ikasleek akatsak egiten dituzte)
- Hiztegi teknikoa edo aurreratuegia
- Akats gramatikal tipikoen gabezia
- Egitura sintaktiko oso erregularrak

Erantzun JSON formatuan:
{{"ia_prob": 0-100, "ia_arrazoiak": ["arrazoia1 euskeraz", "arrazoia2"], "adibideak": ["testu zatia susmagarria 1", "testu zatia 2"]}}

Testua:
{texto[:2000]}"""

            r1 = client.messages.create(
                model='claude-sonnet-4-20250514',
                max_tokens=600,
                messages=[{'role':'user','content': prompt_ia}]
            )

            m1 = re.search(r'\{[^{}]*\}', r1.content[0].text)
            if m1:
                j1 = json.loads(m1.group())
                ia_prob = j1.get('ia_prob', 0)
                ia_arrazoiak = j1.get('ia_arrazoiak', [])
                ia_adibideak = j1.get('adibideak', [])
            else:
                ia_prob = 0
                ia_arrazoiak = []
                ia_adibideak = []

            if ia_prob > 40:
                # IA DETECTADA - No corregir
                print(f'IA:{ia_prob}% - Ez da zuzendu')
                akatsak = 'Ez da zuzendu IA erabilera susmagarria detektatu delako.'
                oharrak = f'IA PROBABILITATEA: {ia_prob}%. ARRAZOIAK: ' + '; '.join(ia_arrazoiak[:3])
                if ia_adibideak:
                    oharrak += f'. ADIBIDE SUSMAGARRIAK: ' + '; '.join([f'"{a[:80]}..."' for a in ia_adibideak[:2]])
                resultados[nombre] = [nombre, 'Idazlan koadernoa', '4B', 0, ia_prob, akatsak, oharrak]
            else:
                # Corregir normalmente
                prompt_corr = f"""Zuzendu euskarazko testu hau (DBH 4, B1 maila MCERL).

EBALUAZIO IRIZPIDEAK:
- Zuzentasun linguistikoa (40%): aditz jokaera, kasu markak, komunztadura
- Edukia eta koherentzia (30%)
- Hiztegia eta adierazpena (20%)
- Originaltasuna (10%)

JSON formatuan erantzun:
{{"nota": 0-10, "akatsak": ["akatsa1 euskeraz", "akatsa2"], "oharrak": "ohar orokorra euskeraz, gutxienez 40 hitz, C2 mailako analisia"}}

Testua:
{texto[:2500]}"""

                r2 = client.messages.create(
                    model='claude-sonnet-4-20250514',
                    max_tokens=800,
                    messages=[{'role':'user','content': prompt_corr}]
                )

                m2 = re.search(r'\{[^{}]*\}', r2.content[0].text)
                if m2:
                    j2 = json.loads(m2.group())
                    nota = j2.get('nota', 5)
                    akatsak = '; '.join(j2.get('akatsak', [])[:4])
                    oharrak = j2.get('oharrak', '')
                    print(f'Nota:{nota}')
                    resultados[nombre] = [nombre, 'Idazlan koadernoa', '4B', nota, ia_prob, akatsak, oharrak]
                else:
                    print('JSON err')
                    resultados[nombre] = [nombre, 'Idazlan koadernoa', '4B', 5, ia_prob, '', '']
        except Exception as e:
            print(f'ERR: {e}')

# Excel
print('\nExcel sortzen...')
wb = Workbook()
ws = wb.active
ws.title = '4B Idazlan Koadernoa'

header_fill = PatternFill('solid', fgColor='4472C4')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ['Ikaslea', 'Aktibitatea', 'Taldea', 'Kalifikazioa', 'IA Prob %', 'Akatsak', 'Ohar Orokorra']

for i, h in enumerate(headers, 1):
    c = ws.cell(1, i, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = header_fill
    c.border = border

datos = sorted(resultados.values(), key=lambda x: x[0])

for r, data in enumerate(datos, 2):
    for c, v in enumerate(data, 1):
        cell = ws.cell(r, c, v)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        if c == 4:  # Nota
            if isinstance(v, (int,float)):
                if v >= 7: cell.fill = PatternFill('solid', fgColor='51CF66')
                elif v >= 5: cell.fill = PatternFill('solid', fgColor='FFE066')
                else: cell.fill = PatternFill('solid', fgColor='FF6B6B')

        if c == 5:  # IA %
            if isinstance(v, (int,float)) and v > 40:
                cell.fill = PatternFill('solid', fgColor='FF6B6B')
                cell.font = Font(bold=True)

widths = [30, 18, 8, 12, 10, 50, 70]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in range(2, len(datos) + 2):
    ws.row_dimensions[row].height = 70

excel_path = r'D:\excel_liburua_idazlana\4B_Idazlan_Koadernoa_FINAL.xlsx'
wb.save(excel_path)
print(f'\nExcel: {excel_path}')
print(f'Total: {len(resultados)} ikasle')
